import pymysql
from openpyxl import *




def fetch_data():
    connection = pymysql.connect(
        host='localhost',
        port=3306,  
        user="root",
        password="caick2510", 
        db='bot_estatistica'
    )

    cursor = connection.cursor(); 

    cursor.execute('select * from Lucro'); 

    rows = cursor.fetchall(); 

    connection.close(); 

    return rows; 

data = fetch_data()


listLucro = [(d[0], d[1], d[-1]) for d in data];  

listMovimento = [(("9hrs as 12hrs: " , d[2]), ("13hrs as 15hrs: " , d[3]), ("16hrs as 18hrs: " , d[4]), ("19hrs as 21hrs: " , d[5]), ("22hrs as 00hrs: " , d[6])) for d in data]

planilha = load_workbook('Relatorio.xlsx'); 

abaLucros = planilha['Vendas Mensais'];
abaMovimento = planilha['Movimento por Hora']; 

for r_idx, (mes, venda, ano) in enumerate(listLucro):
    abaLucros.cell(row=r_idx+5, column=3).value = mes;
    abaLucros.cell(row=r_idx+5, column=4).value = ano;
    abaLucros.cell(row=r_idx+5, column=5).value = venda;

for i, row in enumerate(listMovimento):
    for j, (hora, valor) in enumerate(row):
        if i == 0:
            abaMovimento.cell(row=j+6, column=i+3).value = hora;
            abaMovimento.cell(row=j+6, column=i+4).value = valor;
        else:
            abaMovimento.cell(row=j+6, column=i+4).value = valor;


planilha.save("Relatorio.xlsx")

import win32com.client as win32

outlook = win32.Dispatch('outlook.application')

email = outlook.CreateItem(0); 


email.To = "bertinviana@gmail.com"
email.Subject = "Relatorio Vendas e movimentos"
email.HTMLbody = """
<h3>Relatorio</h3>
<p>Segue o relatorio de Lucros e movimentos deste mes atualizada</p> 
"""

anexo = "C:\\Users\\Caick\\Desktop\\Trabalho_Python\\Relatorio.xlsx"

email.Attachments.Add(anexo); 

email.Send()

print('email enviado'); 